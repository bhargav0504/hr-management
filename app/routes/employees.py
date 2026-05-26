import os
import io
from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, session, send_file, current_app)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app import db
from app.models.employee import Employee
from app.models.family import EmployeeFamily
from app.models.masters import Department, Designation, Category, Location
from flask_wtf import FlaskForm
from wtforms import (StringField, SelectField, DecimalField, DateField,
                     BooleanField, TextAreaField, SubmitField)
from wtforms.validators import DataRequired, Optional, Length, NumberRange
from decimal import Decimal

employees_bp = Blueprint('employees', __name__, url_prefix='/employees')

PHOTO_FOLDER = os.path.join('app', 'static', 'uploads', 'photos')
ALLOWED_PHOTO = {'png', 'jpg', 'jpeg', 'gif', 'webp'}


def _allowed_photo(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_PHOTO


def _get_company_id():
    return session.get('company_id')


def _master_choices(model, company_id):
    items = model.query.filter_by(company_id=company_id, is_active=True).order_by(model.name).all()
    return [('', '— Select —')] + [(i.name, i.name) for i in items]


LEAVING_REASONS = [
    ('', '— Not Applicable —'),
    ('resignation', 'Resignation'),
    ('termination', 'Termination'),
    ('retirement', 'Retirement'),
    ('contract_end', 'Contract End'),
    ('absconding', 'Absconding'),
    ('other', 'Other'),
]


class EmployeeForm(FlaskForm):
    emp_code = StringField('Employee Code *', validators=[DataRequired(), Length(1, 20)])
    first_name = StringField('First Name *', validators=[DataRequired(), Length(1, 100)])
    last_name = StringField('Last Name *', validators=[DataRequired(), Length(1, 100)])
    gender = SelectField('Gender', choices=[('Male', 'Male'), ('Female', 'Female'), ('Transgender', 'Transgender')])
    marital_status = SelectField('Marital Status', choices=[('Single', 'Single'), ('Married', 'Married')])
    father_husband_name = StringField('Father / Husband Name', validators=[Optional(), Length(max=200)])
    date_of_birth = DateField('Date of Birth', validators=[Optional()])
    date_of_joining = DateField('Date of Joining *', validators=[DataRequired()])
    date_of_leaving = DateField('Date of Leaving', validators=[Optional()])
    reason_for_leaving = SelectField('Reason for Leaving', choices=LEAVING_REASONS)
    leaving_remarks = TextAreaField('Leaving Remarks', validators=[Optional()])
    department = SelectField('Department *', choices=[], validators=[DataRequired()])
    location = SelectField('Location *', choices=[], validators=[DataRequired()])
    designation = SelectField('Designation *', choices=[], validators=[DataRequired()])
    category = SelectField('Category', choices=[])
    employment_type = SelectField('Employment Type',
                                  choices=[('permanent', 'Permanent'), ('contract', 'Contract')])
    mobile = StringField('Mobile', validators=[Optional(), Length(max=15)])
    email = StringField('Email', validators=[Optional(), Length(max=120)])
    address = TextAreaField('Address', validators=[Optional()])
    city = StringField('City', validators=[Optional(), Length(max=100)])
    state = StringField('State', validators=[Optional(), Length(max=50)])
    pan_number = StringField('PAN Number', validators=[Optional(), Length(max=10)])
    aadhaar_number = StringField('Aadhaar Number', validators=[Optional(), Length(max=12)])
    uan_number = StringField('UAN Number (PF)', validators=[Optional(), Length(max=12)])
    esic_ip_number = StringField('ESIC IP Number', validators=[Optional(), Length(max=17)])
    bank_account = StringField('Bank Account No.', validators=[Optional(), Length(max=20)])
    bank_name = StringField('Bank Name', validators=[Optional(), Length(max=100)])
    ifsc_code = StringField('IFSC Code', validators=[Optional(), Length(max=11)])
    gross_ctc = DecimalField('Gross CTC', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    basic_salary = DecimalField('Basic Salary *', validators=[DataRequired(), NumberRange(min=0)], places=2)
    hra = DecimalField('HRA', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    da = DecimalField('DA', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    special_allowance = DecimalField('Special Allowance', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    other_allowance = DecimalField('Other Allowance', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    petrol_allowance = DecimalField('Petrol Allowance', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    conveyance = DecimalField('Conveyance', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    medical_allowance = DecimalField('Medical Allowance', validators=[Optional(), NumberRange(min=0)], places=2, default=0)
    pf_applicable = BooleanField('PF Applicable', default=True)
    pf_on_ceiling = BooleanField('PF on ₹15,000 Ceiling', default=True)
    pension_eligible = BooleanField('EPS Eligible', default=True)
    esic_applicable = BooleanField('ESIC Applicable', default=True)
    pt_applicable = BooleanField('PT Applicable', default=True)
    tds_regime = SelectField('TDS Regime', choices=[('new', 'New Regime'), ('old', 'Old Regime')])
    submit = SubmitField('Save Employee')


def _set_master_choices(form, company_id):
    form.department.choices = _master_choices(Department, company_id)
    form.designation.choices = _master_choices(Designation, company_id)
    form.category.choices = [('', '— Select —')] + [(c.name, f'{c.name} ({c.type})') for c in
                              Category.query.filter_by(company_id=company_id, is_active=True).order_by(Category.name).all()]
    locs = Location.query.filter_by(company_id=company_id, is_active=True).order_by(Location.name).all()
    form.location.choices = [('', '— Select Location —')] + [(l.name, l.name) for l in locs]


@employees_bp.route('/')
@login_required
def list_employees():
    cid = _get_company_id()
    show_inactive = request.args.get('inactive', 'false') == 'true'
    q = request.args.get('q', '').strip()
    dept = request.args.get('dept', '')
    query = Employee.query.filter_by(company_id=cid)
    if not show_inactive:
        query = query.filter_by(is_active=True)
    if q:
        query = query.filter(db.or_(
            Employee.first_name.ilike(f'%{q}%'),
            Employee.last_name.ilike(f'%{q}%'),
            Employee.emp_code.ilike(f'%{q}%'),
        ))
    if dept:
        query = query.filter_by(department=dept)
    employees = query.order_by(Employee.emp_code).all()
    departments = [r[0] for r in Employee.query.filter_by(company_id=cid)
                   .with_entities(Employee.department).distinct().all()]
    return render_template('employees/list.html', employees=employees,
                           departments=departments, show_inactive=show_inactive, q=q, dept=dept)


@employees_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_employee():
    cid = _get_company_id()
    form = EmployeeForm()
    _set_master_choices(form, cid)
    if form.validate_on_submit():
        if Employee.query.filter_by(company_id=cid, emp_code=form.emp_code.data).first():
            flash('Employee code already exists in this company.', 'danger')
            return render_template('employees/form.html', form=form, title='Add Employee')
        emp = Employee(company_id=cid)
        _populate_employee(emp, form)
        # Photo
        photo_file = request.files.get('photo')
        if photo_file and photo_file.filename and _allowed_photo(photo_file.filename):
            os.makedirs(PHOTO_FOLDER, exist_ok=True)
            ext = photo_file.filename.rsplit('.', 1)[1].lower()
            filename = secure_filename(f'{form.emp_code.data}.{ext}')
            photo_file.save(os.path.join(PHOTO_FOLDER, filename))
            emp.photo_filename = filename
        db.session.add(emp)
        db.session.commit()
        flash(f'{emp.full_name} added successfully.', 'success')
        return redirect(url_for('employees.detail', emp_id=emp.id))
    return render_template('employees/form.html', form=form, title='Add Employee')


@employees_bp.route('/<int:emp_id>')
@login_required
def detail(emp_id):
    from app.models.payroll import SalaryRecord
    emp = Employee.query.get_or_404(emp_id)
    recent_payroll = emp.salary_records.order_by(
        SalaryRecord.year.desc(), SalaryRecord.month.desc()
    ).limit(12).all()
    active_advances = emp.advances.filter_by(status='active').all()
    return render_template('employees/detail.html', emp=emp,
                           recent_payroll=recent_payroll, active_advances=active_advances)


@employees_bp.route('/<int:emp_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_employee(emp_id):
    cid = _get_company_id()
    emp = Employee.query.get_or_404(emp_id)
    form = EmployeeForm(obj=emp)
    _set_master_choices(form, cid)
    if form.validate_on_submit():
        existing = Employee.query.filter(
            Employee.company_id == cid,
            Employee.emp_code == form.emp_code.data,
            Employee.id != emp_id
        ).first()
        if existing:
            flash('Employee code already used.', 'danger')
        else:
            _populate_employee(emp, form)
            photo_file = request.files.get('photo')
            if photo_file and photo_file.filename and _allowed_photo(photo_file.filename):
                os.makedirs(PHOTO_FOLDER, exist_ok=True)
                ext = photo_file.filename.rsplit('.', 1)[1].lower()
                filename = secure_filename(f'{form.emp_code.data}.{ext}')
                photo_file.save(os.path.join(PHOTO_FOLDER, filename))
                emp.photo_filename = filename
            db.session.commit()
            flash('Employee updated.', 'success')
            return redirect(url_for('employees.detail', emp_id=emp.id))
    return render_template('employees/form.html', form=form, title='Edit Employee', emp=emp)


@employees_bp.route('/<int:emp_id>/delete', methods=['POST'])
@login_required
def delete_employee(emp_id):
    if not current_user.is_admin():
        flash('Only admins can delete employees.', 'danger')
        return redirect(url_for('employees.detail', emp_id=emp_id))
    emp = Employee.query.get_or_404(emp_id)
    from app.models.payroll import SalaryRecord
    if SalaryRecord.query.filter_by(employee_id=emp_id).count() > 0:
        flash('Cannot delete — employee has payroll records. Deactivate instead.', 'warning')
        return redirect(url_for('employees.detail', emp_id=emp_id))
    name = emp.full_name
    db.session.delete(emp)
    db.session.commit()
    flash(f'{name} permanently deleted.', 'success')
    return redirect(url_for('employees.list_employees'))


@employees_bp.route('/<int:emp_id>/toggle', methods=['POST'])
@login_required
def toggle_employee(emp_id):
    if not current_user.is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('employees.list_employees'))
    emp = Employee.query.get_or_404(emp_id)
    emp.is_active = not emp.is_active
    db.session.commit()
    flash(f'{emp.full_name} {"activated" if emp.is_active else "deactivated"}.', 'success')
    return redirect(url_for('employees.detail', emp_id=emp_id))


# ── Family Details ────────────────────────────────────────────────────────────

@employees_bp.route('/<int:emp_id>/family')
@login_required
def family(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    return render_template('employees/family.html', emp=emp)


@employees_bp.route('/<int:emp_id>/family/add', methods=['POST'])
@login_required
def add_family(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    from datetime import datetime as dt
    dob_str = request.form.get('date_of_birth')
    dob = dt.strptime(dob_str, '%Y-%m-%d').date() if dob_str else None
    member = EmployeeFamily(
        employee_id=emp_id,
        name=request.form.get('name', '').strip(),
        relation=request.form.get('relation', ''),
        date_of_birth=dob,
        is_nominee=request.form.get('is_nominee') == 'on',
        nominee_percent=int(request.form.get('nominee_percent', 0) or 0),
    )
    db.session.add(member)
    db.session.commit()
    flash('Family member added.', 'success')
    return redirect(url_for('employees.family', emp_id=emp_id))


@employees_bp.route('/<int:emp_id>/family/<int:member_id>/delete', methods=['POST'])
@login_required
def delete_family(emp_id, member_id):
    member = EmployeeFamily.query.get_or_404(member_id)
    db.session.delete(member)
    db.session.commit()
    flash('Removed.', 'success')
    return redirect(url_for('employees.family', emp_id=emp_id))


# ── Bulk Import ───────────────────────────────────────────────────────────────

@employees_bp.route('/import/template')
@login_required
def import_template():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Import'
    headers = [
        'emp_code*', 'first_name*', 'last_name*', 'gender', 'date_of_birth(YYYY-MM-DD)',
        'date_of_joining*(YYYY-MM-DD)', 'department*', 'designation*', 'category',
        'location', 'employment_type(permanent/contract)', 'mobile', 'email', 'city', 'address',
        'pan_number', 'aadhaar_number', 'uan_number', 'esic_ip_number',
        'bank_name', 'bank_account', 'ifsc_code',
        'basic_salary*', 'hra', 'da', 'special_allowance', 'other_allowance',
        'petrol_allowance', 'conveyance', 'medical_allowance', 'gross_ctc',
        'pf_applicable(YES/NO)', 'esic_applicable(YES/NO)', 'pt_applicable(YES/NO)',
        'tds_regime(new/old)',
    ]
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A56DB')
        cell.column_letter  # touch to set width
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='employee_import_template.xlsx')


@employees_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_employees():
    cid = _get_company_id()
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload a valid .xlsx file.', 'danger')
            return redirect(url_for('employees.import_employees'))
        from openpyxl import load_workbook
        from datetime import datetime as dt
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(c.value).lower().replace('*', '').split('(')[0].strip()
                   for c in ws[1]]
        created = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = dict(zip(headers, row))
            raw_code = data.get('emp_code', '') or ''
            if isinstance(raw_code, float) and raw_code == int(raw_code):
                raw_code = int(raw_code)
            emp_code = str(int(raw_code)).zfill(3) if isinstance(raw_code, int) else str(raw_code).strip()
            if not emp_code:
                continue
            if Employee.query.filter_by(company_id=cid, emp_code=emp_code).first():
                errors.append(f'{emp_code}: already exists, skipped')
                continue
            try:
                def _parse_date(v):
                    if not v:
                        return None
                    if hasattr(v, 'date'):
                        return v.date()
                    return dt.strptime(str(v)[:10], '%Y-%m-%d').date()

                def _yn(v, default=True):
                    if v is None:
                        return default
                    return str(v).strip().upper() not in ('NO', 'N', 'FALSE', '0')

                emp = Employee(
                    company_id=cid,
                    emp_code=emp_code,
                    first_name=str(data.get('first_name', '') or '').strip(),
                    last_name=str(data.get('last_name', '') or '').strip(),
                    gender=str(data.get('gender', '') or 'Male').strip(),
                    date_of_birth=_parse_date(data.get('date_of_birth(yyyy-mm-dd)') or data.get('date_of_birth')),
                    date_of_joining=_parse_date(data.get('date_of_joining*(yyyy-mm-dd)') or data.get('date_of_joining')),
                    department=str(data.get('department', '') or '').strip(),
                    designation=str(data.get('designation', '') or '').strip(),
                    category=str(data.get('category', '') or '').strip(),
                    location=str(data.get('location', '') or 'Head Office').strip(),
                    employment_type=str(data.get('employment_type(permanent/contract)') or data.get('employment_type') or 'permanent').strip(),
                    mobile=str(data.get('mobile', '') or '').strip(),
                    email=str(data.get('email', '') or '').strip(),
                    city=str(data.get('city', '') or '').strip(),
                    address=str(data.get('address', '') or '').strip(),
                    pan_number=str(data.get('pan_number', '') or '').strip().upper(),
                    aadhaar_number=str(data.get('aadhaar_number', '') or '').strip(),
                    uan_number=str(data.get('uan_number', '') or '').strip(),
                    esic_ip_number=str(data.get('esic_ip_number', '') or '').strip(),
                    bank_name=str(data.get('bank_name', '') or '').strip(),
                    bank_account=str(data.get('bank_account', '') or '').strip(),
                    ifsc_code=str(data.get('ifsc_code', '') or '').strip().upper(),
                    basic_salary=Decimal(str(data.get('basic_salary*') or data.get('basic_salary') or 0)),
                    hra=Decimal(str(data.get('hra', 0) or 0)),
                    da=Decimal(str(data.get('da', 0) or 0)),
                    special_allowance=Decimal(str(data.get('special_allowance', 0) or 0)),
                    other_allowance=Decimal(str(data.get('other_allowance', 0) or 0)),
                    petrol_allowance=Decimal(str(data.get('petrol_allowance', 0) or 0)),
                    conveyance=Decimal(str(data.get('conveyance', 0) or 0)),
                    medical_allowance=Decimal(str(data.get('medical_allowance', 0) or 0)),
                    gross_ctc=Decimal(str(data.get('gross_ctc', 0) or 0)),
                    pf_applicable=_yn(data.get('pf_applicable(yes/no)') or data.get('pf_applicable')),
                    esic_applicable=_yn(data.get('esic_applicable(yes/no)') or data.get('esic_applicable')),
                    pt_applicable=_yn(data.get('pt_applicable(yes/no)') or data.get('pt_applicable')),
                    tds_regime=str(data.get('tds_regime(new/old)') or data.get('tds_regime') or 'new').strip(),
                )
                db.session.add(emp)
                created += 1
            except Exception as e:
                errors.append(f'{emp_code}: {e}')
        db.session.commit()
        flash(f'{created} employees imported.', 'success')
        for err in errors:
            flash(err, 'warning')
        return redirect(url_for('employees.list_employees'))
    return render_template('employees/import.html')


def _populate_employee(emp, form):
    for field in [
        'emp_code', 'first_name', 'last_name', 'gender', 'marital_status',
        'father_husband_name', 'date_of_birth', 'date_of_joining', 'date_of_leaving',
        'reason_for_leaving', 'leaving_remarks',
        'department', 'location', 'designation', 'category', 'employment_type',
        'mobile', 'email', 'address', 'city', 'state',
        'pan_number', 'aadhaar_number', 'uan_number', 'esic_ip_number',
        'bank_account', 'bank_name', 'ifsc_code',
        'gross_ctc', 'basic_salary', 'hra', 'da', 'special_allowance', 'other_allowance',
        'petrol_allowance', 'conveyance', 'medical_allowance',
        'pf_applicable', 'pf_on_ceiling', 'pension_eligible',
        'esic_applicable', 'pt_applicable', 'tds_regime',
    ]:
        setattr(emp, field, getattr(form, field).data)
