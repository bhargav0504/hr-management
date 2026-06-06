from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, send_file, current_app, session)
from flask_login import login_required, current_user
from app import db
from app.models.employee import Employee
from app.models.payroll import SalaryRecord, PayrollLock
from app.models.advance import Advance, AdvanceRepayment
from app.utils.payroll_calculator import calculate_payroll
from datetime import date, datetime
import calendar
import io

payroll_bp = Blueprint('payroll', __name__, url_prefix='/payroll')


@payroll_bp.route('/')
@login_required
def index():
    today = date.today()
    company_id = session.get('company_id')
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    records = (
        SalaryRecord.query
        .join(Employee)
        .filter(Employee.company_id == company_id,
                SalaryRecord.month == month, SalaryRecord.year == year)
        .order_by(Employee.emp_code)
        .all()
    )
    employees_with_payroll = {r.employee_id for r in records}
    pending_employees = (
        Employee.query
        .filter(
            Employee.company_id == company_id,
            Employee.is_active == True,
            ~Employee.id.in_(employees_with_payroll)
        )
        .order_by(Employee.emp_code)
        .all()
    )
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(2020, today.year + 2))
    totals = _compute_totals(records)
    lock = PayrollLock.query.filter_by(company_id=company_id, month=month, year=year).first()
    is_locked = lock.is_locked if lock else False
    return render_template('payroll/index.html',
                           records=records, pending_employees=pending_employees,
                           month=month, year=year, months=months, years=years,
                           totals=totals, is_locked=is_locked)


@payroll_bp.route('/run', methods=['GET', 'POST'])
@login_required
def run_payroll():
    """Run payroll for all pending employees in a given month/year."""
    today = date.today()
    company_id = session.get('company_id')
    if request.method == 'POST':
        month = int(request.form.get('month', today.month))
        year = int(request.form.get('year', today.year))
        total_days = int(request.form.get('total_working_days', 26))

        if year > today.year or (year == today.year and month >= today.month):
            flash('Payroll can only be run for months that have already ended.', 'warning')
            return redirect(url_for('payroll.run_payroll', month=month, year=year))

        employees = Employee.query.filter_by(company_id=company_id, is_active=True).all()
        created = 0
        skipped = 0
        for emp in employees:
            existing = SalaryRecord.query.filter_by(
                employee_id=emp.id, month=month, year=year
            ).first()
            if existing:
                skipped += 1
                continue

            # Per-employee overrides from form
            present_days = float(request.form.get(f'present_{emp.id}', total_days))
            extra_incentive = float(request.form.get(f'incentive_{emp.id}', 0) or 0)
            overtime_amt = float(request.form.get(f'overtime_{emp.id}', 0) or 0)
            manual_advance = request.form.get(f'advance_{emp.id}', '').strip()
            tds_override = request.form.get(f'tds_{emp.id}', '').strip()

            advance_deduction = float(manual_advance) if manual_advance else _get_advance_deduction(emp, month, year)
            ytd_tds = _get_ytd_tds(emp, month, year)

            extra = {}
            if extra_incentive: extra['incentive'] = extra_incentive
            if overtime_amt: extra['overtime'] = overtime_amt

            data = calculate_payroll(
                employee=emp,
                month=month,
                year=year,
                present_days=present_days,
                total_working_days=total_days,
                advance_deduction=advance_deduction,
                ytd_tds=ytd_tds,
                config=current_app.config,
                extra_earnings=extra,
            )
            if tds_override:
                data['tds'] = float(tds_override)
            record = SalaryRecord(employee_id=emp.id, **data)
            record.status = 'processed'
            record.processed_at = datetime.utcnow()
            record.processed_by = current_user.id
            db.session.add(record)
            created += 1

        db.session.commit()

        # Auto-lock the month after processing
        lock = PayrollLock.query.filter_by(company_id=company_id, month=month, year=year).first()
        if not lock:
            lock = PayrollLock(company_id=company_id, month=month, year=year)
            db.session.add(lock)
        lock.is_locked = True
        lock.locked_by = current_user.id
        lock.locked_at = datetime.utcnow()
        db.session.commit()

        flash(f'Payroll processed: {created} records created, {skipped} already existed. Month is now locked.', 'success')
        return redirect(url_for('payroll.salary_sheet', month=month, year=year))

    # GET — show form
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    total_days = int(request.args.get('total_days', 26))
    location_filter = request.args.get('location', '')
    from app.models.masters import Location as _Loc
    locations = _Loc.query.filter_by(company_id=company_id, is_active=True).order_by(_Loc.name).all()
    emp_q = Employee.query.filter_by(company_id=company_id, is_active=True)
    if location_filter:
        emp_q = emp_q.filter_by(location=location_filter)
    employees = emp_q.order_by(Employee.emp_code).all()
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(2020, today.year + 2))
    return render_template('payroll/run.html',
                           employees=employees, month=month, year=year,
                           months=months, years=years, total_days=total_days,
                           locations=locations, location_filter=location_filter)


@payroll_bp.route('/record/<int:record_id>')
@login_required
def record_detail(record_id):
    record = SalaryRecord.query.get_or_404(record_id)
    return render_template('payroll/record_detail.html', record=record)


def _check_lock(record):
    """Return flash message and redirect if month is locked, else None."""
    company_id = session.get('company_id')
    lock = PayrollLock.query.filter_by(
        company_id=company_id, month=record.month, year=record.year, is_locked=True
    ).first()
    if lock:
        flash(f'{record.period_label} is locked. Unlock it first to make changes.', 'warning')
        return redirect(url_for('payroll.index', month=record.month, year=record.year))
    return None


@payroll_bp.route('/record/<int:record_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_record(record_id):
    record = SalaryRecord.query.get_or_404(record_id)
    blocked = _check_lock(record)
    if blocked:
        return blocked
    if request.method == 'POST':
        # Manual override fields
        for field in ['present_days', 'tds', 'advance_deduction', 'other_deductions']:
            val = request.form.get(field, 0)
            setattr(record, field, float(val))
        record.other_deductions_remarks = request.form.get('other_deductions_remarks', '')

        # Recalculate using the calculator
        emp = record.employee
        data = calculate_payroll(
            employee=emp,
            month=record.month,
            year=record.year,
            present_days=float(record.present_days),
            total_working_days=record.total_working_days,
            advance_deduction=float(record.advance_deduction),
            ytd_tds=_get_ytd_tds(emp, record.month, record.year, exclude_id=record.id),
            config=current_app.config,
        )
        for k, v in data.items():
            setattr(record, k, v)
        record.tds = float(request.form.get('tds', record.tds))
        record.status = 'processed'
        record.processed_at = datetime.utcnow()
        record.processed_by = current_user.id
        db.session.commit()
        flash('Record updated.', 'success')
        return redirect(url_for('payroll.record_detail', record_id=record.id))
    return render_template('payroll/edit_record.html', record=record)


@payroll_bp.route('/bulk-edit', methods=['GET', 'POST'])
@login_required
def bulk_edit():
    company_id = session.get('company_id')
    today = date.today()
    month = int(request.args.get('month', request.form.get('month', today.month)))
    year = int(request.args.get('year', request.form.get('year', today.year)))

    records = (
        SalaryRecord.query.join(Employee)
        .filter(Employee.company_id == company_id,
                SalaryRecord.month == month, SalaryRecord.year == year)
        .order_by(Employee.emp_code).all()
    )

    lock = PayrollLock.query.filter_by(company_id=company_id, month=month, year=year, is_locked=True).first()
    if lock and request.method == 'POST':
        flash(f'Payroll for this month is locked. Unlock it first.', 'warning')
        return redirect(url_for('payroll.index', month=month, year=year))

    if request.method == 'POST' and request.form.get('action') == 'save':
        updated = 0
        for record in records:
            rid = record.id
            emp = record.employee
            new_days = float(request.form.get(f'days_{rid}', record.present_days))
            new_adv = float(request.form.get(f'adv_{rid}', record.advance_deduction) or 0)
            new_other = float(request.form.get(f'other_{rid}', record.other_deductions) or 0)
            new_other_rem = request.form.get(f'other_rem_{rid}', record.other_deductions_remarks or '')
            new_tds = request.form.get(f'tds_{rid}', '').strip()

            data = calculate_payroll(
                employee=emp,
                month=record.month, year=record.year,
                present_days=new_days,
                total_working_days=record.total_working_days,
                advance_deduction=new_adv,
                other_deductions=new_other,
                other_deductions_remarks=new_other_rem,
                ytd_tds=_get_ytd_tds(emp, record.month, record.year, exclude_id=rid),
                config=current_app.config,
            )
            for k, v in data.items():
                setattr(record, k, v)
            if new_tds:
                record.tds = float(new_tds)
            record.other_deductions_remarks = new_other_rem
            record.status = 'processed'
            record.processed_at = datetime.utcnow()
            record.processed_by = current_user.id
            updated += 1
        db.session.commit()
        flash(f'{updated} records updated.', 'success')
        return redirect(url_for('payroll.salary_sheet', month=month, year=year))

    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(2020, today.year + 2))
    return render_template('payroll/bulk_edit.html', records=records,
                           month=month, year=year, months=months, years=years)


@payroll_bp.route('/lock', methods=['POST'])
@login_required
def lock_month():
    if not current_user.is_admin():
        flash('Only admins can lock payroll months.', 'danger')
        return redirect(request.referrer or url_for('payroll.index'))
    company_id = session.get('company_id')
    month = int(request.form.get('month'))
    year = int(request.form.get('year'))
    lock = PayrollLock.query.filter_by(company_id=company_id, month=month, year=year).first()
    if not lock:
        lock = PayrollLock(company_id=company_id, month=month, year=year)
        db.session.add(lock)
    lock.is_locked = True
    lock.locked_by = current_user.id
    lock.locked_at = datetime.utcnow()
    db.session.commit()
    flash(f'Payroll for {calendar.month_name[month]} {year} locked.', 'success')
    return redirect(url_for('payroll.index', month=month, year=year))


@payroll_bp.route('/unlock', methods=['POST'])
@login_required
def unlock_month():
    if not current_user.is_admin():
        flash('Only admins can unlock payroll months.', 'danger')
        return redirect(request.referrer or url_for('payroll.index'))
    company_id = session.get('company_id')
    month = int(request.form.get('month'))
    year = int(request.form.get('year'))
    password = request.form.get('password', '')
    if not current_user.check_password(password):
        flash('Incorrect password. Payroll remains locked.', 'danger')
        return redirect(url_for('payroll.index', month=month, year=year))
    lock = PayrollLock.query.filter_by(company_id=company_id, month=month, year=year).first()
    if lock:
        lock.is_locked = False
        db.session.commit()
    flash(f'Payroll for {calendar.month_name[month]} {year} unlocked.', 'success')
    return redirect(url_for('payroll.index', month=month, year=year))


@payroll_bp.route('/record/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_record(record_id):
    if not current_user.is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('payroll.index'))
    record = SalaryRecord.query.get_or_404(record_id)
    blocked = _check_lock(record)
    if blocked:
        return blocked
    db.session.delete(record)
    db.session.commit()
    flash('Payroll record deleted.', 'success')
    return redirect(url_for('payroll.index'))


@payroll_bp.route('/record/<int:record_id>/payslip')
@login_required
def download_payslip(record_id):
    from app.utils.pdf_generator import generate_payslip_pdf
    record = SalaryRecord.query.get_or_404(record_id)
    pdf_bytes = generate_payslip_pdf(record, current_app.config)
    buf = io.BytesIO(pdf_bytes)
    buf.seek(0)
    filename = f'payslip_{record.employee.emp_code}_{record.month:02d}_{record.year}.pdf'
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


@payroll_bp.route('/sheet')
@login_required
def salary_sheet():
    today = date.today()
    company_id = session.get('company_id')
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    records = (
        SalaryRecord.query
        .join(Employee)
        .filter(Employee.company_id == company_id,
                SalaryRecord.month == month, SalaryRecord.year == year)
        .order_by(Employee.emp_code)
        .all()
    )
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(2020, today.year + 2))
    totals = _compute_totals(records)
    return render_template('payroll/salary_sheet.html',
                           records=records, month=month, year=year,
                           months=months, years=years, totals=totals)


@payroll_bp.route('/export/excel')
@login_required
def export_excel():
    from app.utils.excel_exporter import export_salary_sheet
    today = date.today()
    company_id = session.get('company_id')
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    records = (
        SalaryRecord.query
        .join(Employee)
        .filter(Employee.company_id == company_id,
                SalaryRecord.month == month, SalaryRecord.year == year)
        .order_by(Employee.emp_code)
        .all()
    )
    wb_bytes = export_salary_sheet(records, month, year, current_app.config)
    buf = io.BytesIO(wb_bytes)
    buf.seek(0)
    filename = f'salary_sheet_{month:02d}_{year}.xlsx'
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@payroll_bp.route('/export/pf-challan')
@login_required
def export_pf_challan():
    from app.utils.excel_exporter import export_pf_challan
    today = date.today()
    company_id = session.get('company_id')
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    records = (
        SalaryRecord.query.join(Employee)
        .filter(Employee.company_id == company_id,
                SalaryRecord.month == month, SalaryRecord.year == year)
        .order_by(Employee.emp_code).all()
    )
    wb_bytes = export_pf_challan(records, month, year, current_app.config)
    buf = io.BytesIO(wb_bytes)
    buf.seek(0)
    filename = f'pf_challan_{month:02d}_{year}.xlsx'
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@payroll_bp.route('/export/esic-challan')
@login_required
def export_esic_challan():
    from app.utils.excel_exporter import export_esic_challan
    today = date.today()
    company_id = session.get('company_id')
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    records = (
        SalaryRecord.query.join(Employee)
        .filter(Employee.company_id == company_id,
                SalaryRecord.month == month, SalaryRecord.year == year)
        .order_by(Employee.emp_code).all()
    )
    wb_bytes = export_esic_challan(records, month, year, current_app.config)
    buf = io.BytesIO(wb_bytes)
    buf.seek(0)
    filename = f'esic_challan_{month:02d}_{year}.xlsx'
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@payroll_bp.route('/export/pt-challan')
@login_required
def export_pt_challan():
    from app.utils.excel_exporter import export_pt_challan
    today = date.today()
    company_id = session.get('company_id')
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    records = (
        SalaryRecord.query.join(Employee)
        .filter(Employee.company_id == company_id,
                SalaryRecord.month == month, SalaryRecord.year == year)
        .order_by(Employee.emp_code).all()
    )
    wb_bytes = export_pt_challan(records, month, year, current_app.config)
    buf = io.BytesIO(wb_bytes)
    buf.seek(0)
    filename = f'pt_challan_{month:02d}_{year}.xlsx'
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@payroll_bp.route('/attendance-template')
@login_required
def attendance_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import session
    company_id = session.get('company_id')
    employees = Employee.query.filter_by(company_id=company_id, is_active=True).order_by(Employee.emp_code).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    headers = ['emp_code', 'present_days', 'performance_bonus', 'incentive', 'overtime_hours', 'overtime_amount', 'other_extra']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='059669')
    for emp in employees:
        ws.append([emp.emp_code, 26, 0, 0, 0, 0, 0])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='attendance_template.xlsx')


@payroll_bp.route('/attendance', methods=['GET', 'POST'])
@login_required
def import_attendance():
    from flask import session
    today = date.today()
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(2020, today.year + 2))
    if request.method == 'POST':
        f = request.files.get('file')
        month = int(request.form.get('month', today.month))
        year = int(request.form.get('year', today.year))
        total_days = int(request.form.get('total_working_days', 26))
        if year > today.year or (year == today.year and month >= today.month):
            flash('Payroll can only be run for months that have already ended.', 'warning')
            return redirect(url_for('payroll.import_attendance'))
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload a valid .xlsx file.', 'danger')
            return redirect(url_for('payroll.import_attendance'))
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in ws[1]]
        company_id = session.get('company_id')
        created = skipped = errors = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = dict(zip(headers, row))
            emp_code = str(data.get('emp_code', '') or '').strip()
            emp = Employee.query.filter_by(company_id=company_id, emp_code=emp_code, is_active=True).first()
            if not emp:
                errors += 1
                continue
            if SalaryRecord.query.filter_by(employee_id=emp.id, month=month, year=year).first():
                skipped += 1
                continue
            present_days = float(data.get('present_days', total_days) or total_days)
            extra = {}
            for key in ['performance_bonus', 'incentive', 'other_extra']:
                val = data.get(key, 0)
                if val:
                    extra[key] = float(val)
            overtime_amt = float(data.get('overtime_amount', 0) or 0)
            if overtime_amt:
                extra['overtime'] = overtime_amt
            advance_deduction = _get_advance_deduction(emp, month, year)
            ytd_tds = _get_ytd_tds(emp, month, year)
            calc_data = calculate_payroll(
                employee=emp, month=month, year=year,
                present_days=present_days, total_working_days=total_days,
                advance_deduction=advance_deduction, ytd_tds=ytd_tds,
                config=current_app.config, extra_earnings=extra,
            )
            record = SalaryRecord(employee_id=emp.id, **calc_data)
            record.status = 'processed'
            record.processed_at = datetime.utcnow()
            record.processed_by = current_user.id
            db.session.add(record)
            created += 1
        db.session.commit()
        _process_advance_repayments(month, year, company_id)
        flash(f'{created} records created, {skipped} skipped, {errors} not found.', 'success')
        return redirect(url_for('payroll.index', month=month, year=year))
    return render_template('payroll/import_attendance.html', months=months, years=years,
                           month=today.month, year=today.year)


def _process_advance_repayments(month, year, company_id):
    records = SalaryRecord.query.join(Employee).filter(
        Employee.company_id == company_id,
        SalaryRecord.month == month,
        SalaryRecord.year == year,
    ).all()
    for record in records:
        if float(record.advance_deduction or 0) > 0:
            emp = record.employee
            active = Advance.query.filter_by(employee_id=emp.id, status='active').all()
            remaining = float(record.advance_deduction)
            for adv in active:
                if remaining <= 0:
                    break
                installment = min(float(adv.monthly_installment), float(adv.balance), remaining)
                if installment > 0:
                    adv.total_repaid = float(adv.total_repaid) + installment
                    adv.balance = float(adv.balance) - installment
                    if adv.balance <= 0:
                        adv.status = 'closed'
                    remaining -= installment
    db.session.commit()


# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_advance_deduction(emp, month, year):
    active_advances = Advance.query.filter_by(employee_id=emp.id, status='active').all()
    total = 0
    for adv in active_advances:
        installment = min(float(adv.monthly_installment), float(adv.balance))
        if installment > 0:
            total += installment
    return total


def _get_ytd_tds(emp, month, year, exclude_id=None):
    fy_start = 4
    if month >= fy_start:
        fy_year = year
    else:
        fy_year = year - 1

    q = SalaryRecord.query.filter(
        SalaryRecord.employee_id == emp.id,
        db.or_(
            db.and_(SalaryRecord.year == fy_year, SalaryRecord.month >= fy_start),
            db.and_(SalaryRecord.year == fy_year + 1, SalaryRecord.month < fy_start),
        )
    )
    if exclude_id:
        q = q.filter(SalaryRecord.id != exclude_id)
    records = q.all()
    return sum(float(r.tds) for r in records)


def _compute_totals(records):
    fields = [
        'gross_earned', 'basic', 'pf_employee', 'pf_employer_total',
        'esic_employee', 'esic_employer', 'pt', 'lwf_employee', 'lwf_employer',
        'tds', 'advance_deduction', 'total_deductions', 'net_salary',
        'gratuity', 'bonus', 'ctc',
    ]
    return {f: sum(float(getattr(r, f) or 0) for r in records) for f in fields}
