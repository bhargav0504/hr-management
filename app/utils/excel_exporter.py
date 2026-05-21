"""
Excel export utilities — layout matches the client's April'26 salary sheet exactly.

Salary sheet column order:
Sr | Code | Name | Gender | DOJ | PF No | ESIC No | Designation | Location |
Gross CTC | Month Days | Present Days | Wage Rate |
[ACTUAL] Basic | HRA | Other Allow |  Gross |
[EARNED] Basic | HRA | Other Allow | Petrol | Salary for PF | Gross |
[DEDUCTIONS] PF | ESIC | PT | TDS | Advance | GLWF | Other | Total Ded |
Net Salary | Bonus |
[EMPLOYER] ER PF | ER ESIC | Gratuity | CTC
"""
import io
import math
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE    = '1a56db'
DKBLUE  = '1e3a8a'
GREEN   = '166534'
RED     = '991b1b'
ORANGE  = '92400e'
LGRAY   = 'f3f4f6'
LBLUE   = 'dbeafe'
LGREEN  = 'dcfce7'
LRED    = 'fee2e2'
LORANGE = 'fef3c7'
WHITE   = 'FFFFFF'


def _cell(ws, row, col, value=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    return c


def _hdr(ws, row, col, value, bg=BLUE, fg=WHITE, bold=True, size=9, wrap=True, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Calibri', bold=bold, color=fg, size=size)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c


def _money(ws, row, col, value, bg=None):
    c = ws.cell(row=row, column=col, value=float(value or 0))
    c.number_format = '#,##0.00'
    c.alignment = Alignment(horizontal='right', vertical='center')
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    return c


def _thin():
    s = Side(style='thin', color='d1d5db')
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_salary_sheet(records, month: int, year: int, config) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]}'{str(year)[-2:]}"

    company = config.get('COMPANY_NAME', 'Company')

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    LAST_COL = 38
    ws.merge_cells(f'A1:{get_column_letter(LAST_COL)}1')
    t = ws['A1']
    t.value = f"{company.upper()} — SALARY SHEET — {calendar.month_name[month].upper()} {year}"
    t.font = Font(name='Calibri', bold=True, size=14, color=DKBLUE)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # ── Row 2: Section headers ─────────────────────────────────────────────────
    # Columns: A-M = employee info, N-Q = ACTUAL salary, R-W = EARNED salary,
    #          X-AE = DEDUCTIONS, AF = Net, AG = Bonus, AH-AK = EMPLOYER, AL = CTC

    # ACTUAL SALARY span: N(14) to Q(17) = 4 cols
    ws.merge_cells('N2:Q2')
    _hdr(ws, 2, 14, 'ACTUAL SALARY', bg=GREEN)

    # EARNED SALARY span: R(18) to W(23) = 6 cols
    ws.merge_cells('R2:W2')
    _hdr(ws, 2, 18, 'EARNED SALARY', bg='0f766e')

    # DEDUCTIONS span: X(24) to AE(31) = 8 cols
    ws.merge_cells('X2:AE2')
    _hdr(ws, 2, 24, 'DEDUCTIONS', bg=RED)

    # EMPLOYER CONTRIBUTIONS: AH(34) to AK(37)
    ws.merge_cells('AH2:AK2')
    _hdr(ws, 2, 34, 'EMPLOYER CONTRIBUTIONS', bg=ORANGE)

    ws.row_dimensions[2].height = 20

    # ── Row 3: Column headers ──────────────────────────────────────────────────
    headers = [
        # A-M: employee info
        ('Sr.', 4), ('Emp Code', 10), ('Name', 22), ('Gender', 7),
        ('DOJ', 11), ('PF No. (UAN)', 14), ('ESIC No.', 14),
        ('Designation', 16), ('Location', 12), ('Gross CTC', 11),
        ('Month\nDays', 7), ('Present\nDays', 7), ('Wage\nRate', 10),
        # N-Q: ACTUAL
        ('Basic', 11), ('HRA', 10), ('Other\nAllow.', 10), ('Gross', 11),
        # R-W: EARNED
        ('Basic', 11), ('HRA', 10), ('Other\nAllow.', 10),
        ('Petrol\nAllow.', 9), ('Salary\nfor PF', 11), ('Gross\nEarned', 11),
        # X-AE: DEDUCTIONS
        ('PF', 10), ('ESIC', 10), ('PT', 8), ('TDS', 10),
        ('Advance', 10), ('GLWF', 8), ('Other\nDed.', 9), ('Total\nDed.', 11),
        # AF-AG
        ('Net\nSalary', 12), ('Bonus', 10),
        # AH-AK: EMPLOYER
        ('ER PF\n(13%)', 11), ('ER ESIC\n(3.25%)', 10),
        ('Gratuity\n(4.81%)', 10), ('CTC', 12),
    ]

    col_bgs = (
        [BLUE]*13 +      # info
        [GREEN]*4 +       # actual
        ['0f766e']*6 +    # earned
        [RED]*8 +         # deductions
        [DKBLUE]*2 +      # net / bonus
        [ORANGE]*4        # employer
    )

    for ci, ((label, width), bg) in enumerate(zip(headers, col_bgs), 1):
        _hdr(ws, 3, ci, label, bg=bg)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[3].height = 34
    ws.freeze_panes = 'A4'

    # ── Data rows ──────────────────────────────────────────────────────────────
    for ri, r in enumerate(records, 1):
        row = ri + 3
        emp = r.employee
        bg = WHITE if ri % 2 == 0 else LGRAY

        def tc(col, val, is_money=False, cell_bg=None):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill('solid', fgColor=cell_bg or bg)
            c.border = _thin()
            c.font = Font(name='Calibri', size=9)
            if is_money:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.alignment = Alignment(vertical='center', wrap_text=False)
            return c

        def tm(col, val, cell_bg=None):
            return tc(col, float(val or 0), is_money=True, cell_bg=cell_bg)

        # Employee info
        tc(1, ri)
        tc(2, emp.emp_code)
        tc(3, emp.full_name)
        tc(4, getattr(emp, 'gender', 'Male'))
        doj = emp.date_of_joining
        c = ws.cell(row=row, column=5, value=doj)
        c.number_format = 'DD-MMM-YY'
        c.fill = PatternFill('solid', fgColor=bg)
        c.border = _thin()
        c.font = Font(name='Calibri', size=9)
        tc(6, emp.uan_number or 'NA')
        tc(7, emp.esic_ip_number or 'NA')
        tc(8, emp.designation)
        tc(9, getattr(emp, 'location', 'Head Office') or 'Head Office')
        tm(10, r.gross_earned)  # Gross CTC (earned this month)
        tc(11, r.total_working_days)
        c2 = ws.cell(row=row, column=12, value=float(r.present_days))
        c2.fill = PatternFill('solid', fgColor=bg); c2.border = _thin()
        c2.font = Font(name='Calibri', size=9)
        c2.alignment = Alignment(horizontal='center', vertical='center')

        # Wage rate = basic / 26
        wage_rate = float(r.basic or 0) / 26 if float(r.basic or 0) > 0 else 0
        tm(13, round(wage_rate, 2))

        # ACTUAL salary (full month)
        tm(14, emp.basic_salary, LGREEN)
        tm(15, emp.hra, LGREEN)
        actual_other = float(emp.gross_salary) - float(emp.basic_salary) - float(emp.hra)
        tm(16, round(actual_other, 2), LGREEN)
        tm(17, emp.gross_salary, LGREEN)

        # EARNED salary (prorated)
        tm(18, r.basic, LBLUE)
        tm(19, r.hra, LBLUE)
        earned_other = float(r.other_allowance or 0) + float(r.da or 0) + float(r.special_allowance or 0) + float(r.conveyance or 0) + float(r.medical_allowance or 0)
        tm(20, round(earned_other, 2), LBLUE)
        tm(21, r.petrol_allowance, LBLUE)
        tm(22, r.pf_wage_base, LBLUE)
        tm(23, r.gross_earned, LBLUE)

        # DEDUCTIONS
        tm(24, r.pf_employee, LRED)
        tm(25, r.esic_employee, LRED)
        tm(26, r.pt, LRED)
        tm(27, r.tds, LRED)
        tm(28, r.advance_deduction, LRED)
        tm(29, r.lwf_employee, LRED)
        tm(30, r.other_deductions, LRED)
        tm(31, r.total_deductions, LRED)

        # Net & Bonus
        net_c = tm(32, r.net_salary)
        net_c.font = Font(name='Calibri', bold=True, size=9, color='166534')
        tm(33, r.bonus)

        # Employer
        tm(34, r.pf_employer_total, LORANGE)
        tm(35, r.esic_employer, LORANGE)
        tm(36, r.gratuity, LORANGE)
        ctc_c = tm(37, r.ctc, LORANGE)
        ctc_c.font = Font(name='Calibri', bold=True, size=9)

    # ── Totals row ─────────────────────────────────────────────────────────────
    tr = len(records) + 4
    ws.merge_cells(f'A{tr}:M{tr}')
    tot_lbl = ws.cell(row=tr, column=1, value='TOTALS')
    tot_lbl.font = Font(bold=True, color=WHITE, size=10)
    tot_lbl.fill = PatternFill('solid', fgColor=DKBLUE)
    tot_lbl.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[tr].height = 18

    money_cols = list(range(10, 14)) + list(range(14, 38))
    for ci in money_cols:
        vals = []
        for ri in range(4, 4 + len(records)):
            v = ws.cell(row=ri, column=ci).value
            if isinstance(v, (int, float)):
                vals.append(v)
        total = sum(vals)
        c = ws.cell(row=tr, column=ci, value=total)
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill('solid', fgColor=DKBLUE)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = _thin()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pf_challan(records, month: int, year: int, config) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f'PF {calendar.month_abbr[month]}{year}'

    company = config.get('COMPANY_NAME', 'Company')
    pf_no = config.get('COMPANY_PF_NO', '')

    ws.merge_cells('A1:K1')
    ws['A1'] = f'PF CHALLAN — {company.upper()} — {calendar.month_name[month].upper()} {year}'
    ws['A1'].font = Font(bold=True, size=13, color=DKBLUE)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 22

    if pf_no:
        ws.merge_cells('A2:K2')
        ws['A2'] = f'PF Registration No.: {pf_no}'
        ws['A2'].alignment = Alignment(horizontal='center')

    hdrs = [
        ('Sr.', 5), ('UAN', 15), ('Employee Name', 24), ('Emp Code', 10),
        ('PF Wages\n(₹)', 12), ('Employee PF\n12% (₹)', 13),
        ('Employer EPF\n3.67% (₹)', 14), ('Employer EPS\n8.33% (₹)', 14),
        ('EDLI\n0.5% (₹)', 11), ('Total ER\nContrib. (₹)', 14), ('Total\nChallan (₹)', 14),
    ]
    for ci, (h, w) in enumerate(hdrs, 1):
        _hdr(ws, 3, ci, h, bg=DKBLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 30

    pf_recs = [r for r in records if r.employee.pf_applicable]
    for ri, r in enumerate(pf_recs, 1):
        row = ri + 3
        emp = r.employee
        bg = WHITE if ri % 2 == 0 else LGRAY
        ee_pf = float(r.pf_employee or 0)
        er_epf = float(r.pf_employer_epf or 0)
        er_eps = float(r.pf_employer_eps or 0)
        er_edli = float(r.pf_employer_edli or 0)
        er_total = float(r.pf_employer_total or 0)
        row_data = [ri, emp.uan_number or '', emp.full_name, emp.emp_code,
                    float(r.pf_wage_base or 0), ee_pf, er_epf, er_eps,
                    er_edli, er_total, ee_pf + er_total]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = PatternFill('solid', fgColor=bg)
            c.border = _thin()
            c.font = Font(name='Calibri', size=9)
            if ci >= 5:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    # Totals
    tr = len(pf_recs) + 4
    ws.merge_cells(f'A{tr}:D{tr}')
    tc = ws.cell(row=tr, column=1, value='TOTAL')
    tc.font = Font(bold=True, color=WHITE)
    tc.fill = PatternFill('solid', fgColor=DKBLUE)
    tc.alignment = Alignment(horizontal='center')
    for ci in range(5, 12):
        vals = [ws.cell(row=ri+3, column=ci).value or 0 for ri in range(1, len(pf_recs)+1)]
        c = ws.cell(row=tr, column=ci, value=sum(vals))
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=DKBLUE)
        c.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_esic_challan(records, month: int, year: int, config) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f'ESIC {calendar.month_abbr[month]}{year}'

    company = config.get('COMPANY_NAME', 'Company')
    esic_no = config.get('COMPANY_ESIC_NO', '')

    ws.merge_cells('A1:H1')
    ws['A1'] = f'ESIC CHALLAN — {company.upper()} — {calendar.month_name[month].upper()} {year}'
    ws['A1'].font = Font(bold=True, size=13, color=DKBLUE)
    ws['A1'].alignment = Alignment(horizontal='center')

    if esic_no:
        ws.merge_cells('A2:H2')
        ws['A2'] = f'ESIC Reg. No.: {esic_no}'
        ws['A2'].alignment = Alignment(horizontal='center')

    hdrs = [
        ('Sr.', 5), ('ESIC IP No.', 18), ('Employee Name', 24), ('Emp Code', 10),
        ('Gross Wages (₹)', 14), ('Employee 0.75% (₹)', 16),
        ('Employer 3.25% (₹)', 16), ('Total ESIC (₹)', 13),
    ]
    for ci, (h, w) in enumerate(hdrs, 1):
        _hdr(ws, 3, ci, h, bg=DKBLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 30

    esic_recs = [r for r in records if float(r.esic_employee or 0) > 0]
    for ri, r in enumerate(esic_recs, 1):
        row = ri + 3
        emp = r.employee
        bg = WHITE if ri % 2 == 0 else LGRAY
        ee = float(r.esic_employee or 0)
        er = float(r.esic_employer or 0)
        row_data = [ri, emp.esic_ip_number or '', emp.full_name, emp.emp_code,
                    float(r.gross_earned or 0), ee, er, ee + er]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = PatternFill('solid', fgColor=bg)
            c.border = _thin()
            c.font = Font(name='Calibri', size=9)
            if ci >= 5:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    tr = len(esic_recs) + 4
    ws.merge_cells(f'A{tr}:D{tr}')
    tc = ws.cell(row=tr, column=1, value='TOTAL')
    tc.font = Font(bold=True, color=WHITE)
    tc.fill = PatternFill('solid', fgColor=DKBLUE)
    tc.alignment = Alignment(horizontal='center')
    for ci in range(5, 9):
        vals = [ws.cell(row=ri+3, column=ci).value or 0 for ri in range(1, len(esic_recs)+1)]
        c = ws.cell(row=tr, column=ci, value=sum(vals))
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=DKBLUE)
        c.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pt_challan(records, month: int, year: int, config) -> bytes:
    """Gujarat Professional Tax challan."""
    wb = Workbook()
    ws = wb.active
    ws.title = f'PT {calendar.month_abbr[month]}{year}'

    company = config.get('COMPANY_NAME', 'Company')
    pt_no = config.get('COMPANY_PT_NO', '')

    ws.merge_cells('A1:G1')
    ws['A1'] = f'PROFESSIONAL TAX CHALLAN — {company.upper()} — {calendar.month_name[month].upper()} {year}'
    ws['A1'].font = Font(bold=True, size=13, color=DKBLUE)
    ws['A1'].alignment = Alignment(horizontal='center')

    if pt_no:
        ws.merge_cells('A2:G2')
        ws['A2'] = f'PT Registration No.: {pt_no}'
        ws['A2'].alignment = Alignment(horizontal='center')

    hdrs = [('Sr.', 5), ('Emp Code', 10), ('Employee Name', 24),
            ('Designation', 16), ('Gross Salary (₹)', 14), ('PT Deducted (₹)', 14), ('Remarks', 14)]
    for ci, (h, w) in enumerate(hdrs, 1):
        _hdr(ws, 3, ci, h, bg=DKBLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 24

    pt_recs = [r for r in records if float(r.pt or 0) > 0]
    for ri, r in enumerate(pt_recs, 1):
        row = ri + 3
        emp = r.employee
        bg = WHITE if ri % 2 == 0 else LGRAY
        row_data = [ri, emp.emp_code, emp.full_name, emp.designation,
                    float(r.gross_earned or 0), float(r.pt or 0),
                    'Gujarat PT — Gross ≥ ₹12,000']
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = PatternFill('solid', fgColor=bg)
            c.border = _thin()
            c.font = Font(name='Calibri', size=9)
            if ci in (5, 6):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    tr = len(pt_recs) + 4
    ws.merge_cells(f'A{tr}:E{tr}')
    tc = ws.cell(row=tr, column=1, value=f'TOTAL PT — {len(pt_recs)} employees')
    tc.font = Font(bold=True, color=WHITE)
    tc.fill = PatternFill('solid', fgColor=DKBLUE)
    tc.alignment = Alignment(horizontal='center')
    total_pt = sum(float(r.pt or 0) for r in pt_recs)
    c = ws.cell(row=tr, column=6, value=total_pt)
    c.number_format = '#,##0.00'
    c.font = Font(bold=True, color=WHITE)
    c.fill = PatternFill('solid', fgColor=DKBLUE)
    c.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_tds_summary_excel(records, fy_year: int, config) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f'TDS FY{fy_year}-{fy_year+1}'

    company = config.get('COMPANY_NAME', 'Company')
    tan = config.get('COMPANY_TAN', '')

    ws.merge_cells('A1:S1')
    ws['A1'] = f'TDS SUMMARY — {company.upper()} — FY {fy_year}-{fy_year+1}'
    ws['A1'].font = Font(bold=True, size=13, color=DKBLUE)
    ws['A1'].alignment = Alignment(horizontal='center')
    if tan:
        ws.merge_cells('A2:S2')
        ws['A2'] = f'TAN: {tan}'
        ws['A2'].alignment = Alignment(horizontal='center')

    month_order = list(range(4, 13)) + list(range(1, 4))
    month_labels = [calendar.month_abbr[m] for m in month_order]

    headers = ['Emp Code', 'Name', 'PAN', 'Designation', 'Regime'] + month_labels + ['Total TDS (₹)', 'Annual Gross (₹)']
    col_widths = [10, 22, 12, 16, 8] + [8]*12 + [13, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws, 3, ci, h, bg=DKBLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 24

    emp_dict = {}
    for r in records:
        eid = r.employee_id
        if eid not in emp_dict:
            emp_dict[eid] = {'emp': r.employee, 'months': {}, 'gross': 0}
        emp_dict[eid]['months'][r.month] = float(r.tds or 0)
        emp_dict[eid]['gross'] += float(r.gross_earned or 0)

    for ri, (_, data) in enumerate(emp_dict.items(), 1):
        emp = data['emp']
        row = ri + 3
        bg = WHITE if ri % 2 == 0 else LGRAY
        month_tds = [data['months'].get(m, 0) for m in month_order]
        total_tds = sum(month_tds)
        row_data = [emp.emp_code, emp.full_name, emp.pan_number or '',
                    emp.designation, (emp.tds_regime or 'new').upper()] + month_tds + [total_tds, data['gross']]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = PatternFill('solid', fgColor=bg)
            c.border = _thin()
            c.font = Font(name='Calibri', size=9)
            if ci >= 6:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
